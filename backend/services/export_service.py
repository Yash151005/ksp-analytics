"""
Export service for KSP Analytics Platform
Handles CSV, PDF, and Excel exports
"""
import csv
import json
from io import StringIO, BytesIO
from datetime import datetime
from typing import List, Dict, Any
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors


def export_crimes_to_csv(crimes: List[Dict[str, Any]]) -> str:
    """Export crime records to CSV format"""
    output = StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[
            "Crime_ID", "Date", "Time", "Type", "Subtype", "District",
            "Taluk", "Police_Station", "Latitude", "Longitude", "Status",
            "Severity", "Weather", "Time_of_Day", "Description"
        ]
    )
    
    writer.writeheader()
    for crime in crimes:
        date_str = crime.get("date")
        if isinstance(date_str, datetime):
            date_str = date_str.strftime("%Y-%m-%d")
        
        lat = crime.get("latitude")
        lng = crime.get("longitude")
        lat_str = f"{lat:.6f}" if lat is not None else ""
        lng_str = f"{lng:.6f}" if lng is not None else ""
        
        writer.writerow({
            "Crime_ID": crime.get("crime_id", ""),
            "Date": date_str,
            "Time": crime.get("time", ""),
            "Type": crime.get("type", ""),
            "Subtype": crime.get("subtype", ""),
            "District": crime.get("district", ""),
            "Taluk": crime.get("taluk", ""),
            "Police_Station": crime.get("police_station", ""),
            "Latitude": lat_str,
            "Longitude": lng_str,
            "Status": crime.get("status", ""),
            "Severity": crime.get("severity", ""),
            "Weather": crime.get("weather", ""),
            "Time_of_Day": crime.get("time_of_day", ""),
            "Description": crime.get("description", ""),
        })
    
    return output.getvalue()


def export_criminals_to_csv(criminals: List[Dict[str, Any]]) -> str:
    """Export criminal records to CSV format"""
    output = StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[
            "Name", "Alias", "Age", "Gender", "Status", "Risk_Score",
            "Crime_Count", "Last_Known_Location", "Known_Associates"
        ]
    )
    
    writer.writeheader()
    for criminal in criminals:
        risk_score = criminal.get("risk_score")
        rs_str = f"{risk_score:.2f}" if risk_score is not None else ""
        
        writer.writerow({
            "Name": criminal.get("name", ""),
            "Alias": criminal.get("alias", ""),
            "Age": criminal.get("age", ""),
            "Gender": criminal.get("gender", ""),
            "Status": criminal.get("status", ""),
            "Risk_Score": rs_str,
            "Crime_Count": len(criminal.get("crime_history", [])),
            "Last_Known_Location": criminal.get("last_known_location", ""),
            "Known_Associates": len(criminal.get("associated_criminal_ids", [])),
        })
    
    return output.getvalue()


def export_report_to_pdf(
    title: str,
    date_range: str,
    summary: Dict[str, Any],
    narrative: str,
    crime_data: List[Dict[str, Any]] = None,
) -> bytes:
    """Generate professional PDF report"""
    pdf_buffer = BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0A1628'),
        spaceAfter=12,
        alignment=1,  # center
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1E3A5F'),
        spaceAfter=10,
        spaceBefore=10,
    )
    
    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['BodyText'],
        fontSize=10,
        leading=14,
    )
    
    # Title
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"Report Period: {date_range}", styles['Normal']))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary statistics
    elements.append(Paragraph("Executive Summary", heading_style))
    
    summary_data = [
        ["Metric", "Value"],
    ]
    
    if isinstance(summary, dict):
        for key, value in summary.items():
            summary_data.append([key.replace("_", " ").title(), str(value)])
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Narrative
    if narrative:
        elements.append(Paragraph("Analysis & Recommendations", heading_style))
        elements.append(Paragraph(narrative, body_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Crime data table if provided
    if crime_data:
        elements.append(PageBreak())
        elements.append(Paragraph("Detailed Crime Records", heading_style))
        
        crime_table_data = [
            ["Crime ID", "Type", "District", "Date", "Status", "Severity"],
        ]
        
        for crime in crime_data[:20]:  # Limit to 20 records per page
            crime_table_data.append([
                crime.get("crime_id", ""),
                crime.get("type", ""),
                crime.get("district", ""),
                crime.get("date", ""),
                crime.get("status", ""),
                str(crime.get("severity", "")),
            ])
        
        crime_table = Table(crime_table_data, colWidths=[1.2*inch, 1*inch, 1.2*inch, 1*inch, 0.8*inch, 0.8*inch])
        crime_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(crime_table)
    
    # Build PDF
    doc.build(elements)
    return pdf_buffer.getvalue()


def export_alerts_to_csv(alerts: List[Dict[str, Any]]) -> str:
    """Export alert records to CSV format"""
    output = StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[
            "Title", "Severity", "Affected_Area", "Crime_Count_Spike",
            "Created_At", "Is_Acknowledged", "Recommendation"
        ]
    )
    
    writer.writeheader()
    for alert in alerts:
        writer.writerow({
            "Title": alert.get("title", ""),
            "Severity": alert.get("severity", ""),
            "Affected_Area": alert.get("affected_area", ""),
            "Crime_Count_Spike": alert.get("crime_count_spike", ""),
            "Created_At": alert.get("created_at", ""),
            "Is_Acknowledged": alert.get("is_acknowledged", ""),
            "Recommendation": alert.get("recommendation", ""),
        })
    
    return output.getvalue()


def export_analytics_to_json(analytics_data: Dict[str, Any]) -> str:
    """Export analytics data to JSON format"""
    return json.dumps(analytics_data, indent=2, default=str)


def create_excel_report(title: str, data_sheets: Dict[str, List[Dict]]) -> bytes:
    """
    Create Excel report with multiple sheets
    Requires openpyxl library
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for sheet_name, data in data_sheets.items():
            ws = wb.create_sheet(sheet_name)
            
            if not data:
                continue
            
            # Add header
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
        
        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    except ImportError:
        # Fallback: return JSON if openpyxl not available
        return json.dumps(data_sheets, indent=2, default=str).encode()
